import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# === НАСТРОЙКИ ===
VARIANT = {
    "function": lambda x: (x - 12) ** 2,
    "a": 5,
    "b": 20,
    "true_min_x": 12,
}

EPSILONS = [10**(-i) for i in range(1, 2)]  # [0.1, 0.01, ..., 1e-7]
EXCEL_FILE = "lab1_excel.xlsx"
PLOT_FILE = "temp_plot.png"

def metod_dihotomii(f, a, b, eps):
    delta = eps / 3.0
    shagi = []
    chislo_vychisleniy = 0
    iteratsiya = 1
    while b - a > eps:
        mid = (a + b) / 2
        x1, x2 = mid - delta, mid + delta
        f1, f2 = f(x1), f(x2)
        chislo_vychisleniy += 2
        shagi.append({"i": iteratsiya, "x1": x1, "x2": x2, "f(x1)": f1, "f(x2)": f2, "a_i": a, "b_i": b, "b_i - a_i": b - a})
        if f1 < f2:
            b = x2
        else:
            a = x1
        iteratsiya += 1
    return (a + b) / 2, chislo_vychisleniy, shagi


def metod_zolotogo_secheniya(f, a, b, eps):
    k = 2 - (1 + np.sqrt(5)) / 2
    x1 = a + k * (b - a)
    x2 = b - k * (b - a)
    f1, f2 = f(x1), f(x2)
    chislo_vychisleniy = 2
    shagi = []
    iteratsiya = 1
    while b - a > eps:
        shagi.append({"i": iteratsiya, "x1": x1, "x2": x2, "f(x1)": f1, "f(x2)": f2, "a_i": a, "b_i": b, "b_i - a_i": b - a})
        if f1 < f2:
            b, x2, f2 = x2, x1, f1
            x1 = a + k * (b - a)
            f1 = f(x1)
            chislo_vychisleniy += 1
        else:
            a, x1, f1 = x1, x2, f2
            x2 = b - k * (b - a)
            f2 = f(x2)
            chislo_vychisleniy += 1
        iteratsiya += 1
    return (a + b) / 2, chislo_vychisleniy, shagi


def dobavit_koef_szhim(df):
    dliny = df["b_i - a_i"].tolist()
    koef = [None]
    for i in range(1, len(dliny)):
        pred, tek = dliny[i-1], dliny[i]
        koef.append(pred / tek if tek != 0 else np.nan)
    df["Коэф. сжатия"] = koef
    return df


def format_eps(eps):
    """Преобразует 1e-3 → '0.001', 1e-1 → '0.1' и т.д."""
    s = f"{eps:.10f}".rstrip('0').rstrip('.')
    return s if '.' in s else s + '.0'


# === ОСНОВНАЯ ФУНКЦИЯ ===
def main():
    f = VARIANT["function"]
    a0, b0 = VARIANT["a"], VARIANT["b"]
    istinnyy_min = VARIANT["true_min_x"]
    demo_eps = 1e-1

    # === Сбор данных ===
    sravnenie = []
    shagi_dih = shagi_zol = None

    for eps in EPSILONS:
        x_d, v_d, s_d = metod_dihotomii(f, a0, b0, eps)
        x_g, v_g, s_g = metod_zolotogo_secheniya(f, a0, b0, eps)
        
        sravnenie.append({
            "ε": format_eps(eps),
            "Выч. (дихотомия)": v_d,
            "Выч. (золотое сечение)": v_g
        })
        if abs(eps - demo_eps) < 1e-10:
            shagi_dih, shagi_zol = s_d, s_g

    df_srav = pd.DataFrame(sravnenie)

    # === ВЫВОД В КОНСОЛЬ ===
    print("=" * 70)
    print("РЕЗУЛЬТАТЫ ОПТИМИЗАЦИИ")
    print("=" * 70)
    print(f"Функция: f(x) = (x - {istinnyy_min})²")
    print(f"Отрезок: [{a0}, {b0}]")
    print(f"Истинный минимум: x = {istinnyy_min}")
    print("\nСРАВНЕНИЕ:")
    print(df_srav.to_string(index=False))

    if shagi_dih is not None:
        print("\n" + "-"*60)
        print(f"ДЕТАЛЬНЫЕ ШАГИ: ДИХОТОМИЯ (ε = {format_eps(demo_eps)})")
        print("-"*60)
        df_d = pd.DataFrame(shagi_dih)
        df_d = dobavit_koef_szhim(df_d)
        print(df_d.to_string(index=False, float_format="%.6g"))

    if shagi_zol is not None:
        print("\n" + "-"*60)
        print(f"ДЕТАЛЬНЫЕ ШАГИ: ЗОЛОТОЕ СЕЧЕНИЕ (ε = {format_eps(demo_eps)})")
        print("-"*60)
        df_g = pd.DataFrame(shagi_zol)
        df_g = dobavit_koef_szhim(df_g)
        print(df_g.to_string(index=False, float_format="%.6g"))

    print("\n" + "" * 30)
    print("ВЫВОД:")
    print("" * 30)
    print("Метод золотого сечения эффективнее дихотомии: он требует меньше вычислений")
    print("функции при той же точности за счёт повторного использования значений.")
    print("Рекомендуется использовать метод золотого сечения для унимодальных функций.")

    # === СОХРАНЕНИЕ В EXCEL ===
    plt.figure(figsize=(9, 5))
    plt.loglog(EPSILONS, df_srav["Выч. (дихотомия)"], 'o-', label="Дихотомия")
    plt.loglog(EPSILONS, df_srav["Выч. (золотое сечение)"], 's-', label="Золотое сечение")
    plt.gca().invert_xaxis()
    plt.xlabel("Точность (ε)")
    plt.ylabel("Число вычислений функции")
    plt.title("Сравнение эффективности методов")
    plt.grid(True, which="both", ls="--", alpha=0.7)
    plt.legend()
    plt.tight_layout()
    plt.savefig(PLOT_FILE, dpi=200)
    plt.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение и вывод"

    for r in dataframe_to_rows(df_srav, index=False, header=True):
        ws.append(r)

    ws.append([])
    ws.append(["ВЫВОД (ЗАКЛЮЧЕНИЕ):"])
    ws.append([
        "Метод золотого сечения эффективнее метода дихотомии. "
        "Он требует меньше вычислений функции при одинаковой точности, "
        "поскольку на каждом шаге переиспользует одно значение функции "
        "из предыдущей итерации. Для поиска минимума унимодальной функции "
        "рекомендуется использовать метод золотого сечения."
    ])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 80

    if shagi_dih:
        ws_d = wb.create_sheet("Дихотомия_1e-4")
        df_d = pd.DataFrame(shagi_dih)
        df_d = dobavit_koef_szhim(df_d)
        for r in dataframe_to_rows(df_d, index=False, header=True):
            ws_d.append(r)

    if shagi_zol:
        ws_g = wb.create_sheet("Золотое_сечение_1e-4")
        df_g = pd.DataFrame(shagi_zol)
        df_g = dobavit_koef_szhim(df_g)
        for r in dataframe_to_rows(df_g, index=False, header=True):
            ws_g.append(r)

    img = ExcelImage(PLOT_FILE)
    ws_plot = wb.create_sheet("График")
    ws_plot.add_image(img, 'A1')

    wb.save(EXCEL_FILE)

    if os.path.exists(PLOT_FILE):
        os.remove(PLOT_FILE)

    print(f"\n Все данные сохранены в Excel: {os.path.abspath(EXCEL_FILE)}")


if __name__ == "__main__":
    main()